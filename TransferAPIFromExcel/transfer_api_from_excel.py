# -*- coding: utf-8 -*-
import os
import time

import openpyxl
import json
import re

# @Date:Jan 2019
# @Author: Enoch
# @Description: this tool is to transfer excel file to API test cases, now we just do data.
# @Usage: put .xlsx file in the same path with .py file and run:
# python3 transfer_api_from_excel.py


class TransferAPIFromExcel():
    def __init__(self, ws_name, wb_name, title_col):
        self.wb = openpyxl.load_workbook(ws_name)
        self.ws = self.wb[wb_name]
        self.col_length = self.get_col_length_by_col_id(title_col)

    def get_col_id_by_name(self, col_name):
        """
        input column name, i.e. 'Expected Result', to output col id for further operations
        """
        for index in range(1, 100):
            row = self.ws.cell(row=1, column=index)
            if row.value and col_name in row.value:
                return index
        return False

    def strip_response_data_by_col_name(self, col_name):
        """
        after get col id, in this funtion we will strip response data, and now using re is better

        response is a 2-dimension list with code list and data list inside
        """
        col_id = self.get_col_id_by_name(col_name)
        # print (col_id)
        code_list = []
        data_list = []
        for i in range(1, self.col_length):
            raw = self.ws.cell(row=i, column=col_id)
            if not raw.value:
                raw.value = "code 0 {'[err_1]': 'response code empty'}"
            r = raw.value.strip().lower()
            # print(r)
            # print ("=====raw print finish====")

            # Try to strip response code
            # here is to find start/stop point
            regex = r"[0-9]{3,5}"
            m = re.search(regex, r)
            if m:
                code = m.group(0)
            else:
                code = 0
            code_list.append(code)

            # dealing with data now
            r_data_start = r.find('{')
            r_data_end = r.find('}')
            data = r[r_data_start:r_data_end+1].strip()

            if not self.get_is_json(data):
                print("[err_3]: response data format err")
                data = {"[err]": "response data format err"}

            data_list.append(data)

        response_data = [code_list, data_list]
        return response_data

    def fill_in_template_response(self,input_list,file_name):
        """
        put the needed data into a test-data template

        for now is ok, maybe will make this common function later
        """
        # init, devide code and response body
        code_list=input_list[0]
        response_body_list=input_list[1]
        content_list=[]

        # generate the needed
        # start with line 2, since 0, and 1 are empty
        for i in range(2,len(code_list)):
            cell=[]
            #  read a template file repeatedly, marked TODO
            f = open(file_name, 'r')
            for line in f:
                if "statusCode" in line:
                    if code_list[i] is "":
                        code_list[i] = "0"

                    line = line.replace("statusCode", '"statusCode"'+': '+ str(code_list[i]))

                if "responseBody" in line:
                    if '{' not in response_body_list[i]:
                        response_body_list[i] = '{}'
                    line = line.replace("responseBody",
                                        '"responseBody"' + ': ' + response_body_list[i])
                cell.append(line)

            f.close()
            content_list.append(cell)
        for x in content_list:
            print(x)
        return content_list

    def get_json_pretty_print(self, json_object):
        """
        just try to make json looks pretty
        """
        return json.dumps(json_object, sort_keys=True, indent=4, separators=(',', ': '))

    def json_encode(self, input_list):
        """
        input a list, transfer it to str, and do format operation.

        and then transfer it to json,

        and then make it pretty.

        for now i will check if it is not "json" enough, just output and err json.
        """

        input_str = ''.join(input_list)
        input_str = self.str_custom_strip(input_str)

        # often error here
        if self.get_is_json(input_str):
            js = json.loads(input_str)
            js_output = json.dumps(js, sort_keys=False, indent=4)
        else:
            js_output = "{[err]: data is not json enough, format err", "data as follows: "+input_str+"}"
            js_output = self.str_custom_strip(str(js_output))

        return js_output


    def str_encode_simple(self, input_list):
        """
        this is a simple version of json_encode

        just input a list, which is kind of json

        do some operation and make it str
        """
        input_str = ''.join(input_list)
        input_str = self.str_custom_strip(str(input_str))
        return input_str

    def output_to_file(self, title_list, content_list, file_name):
        """
        output to a file, ideally it is formatted

        """
        f = open(file_name,'w')
 
        for idx, cell in enumerate(content_list):
            print("output case number: "+str(idx+1))

            # format single test data and output to file
            formatted_cell=self.json_encode(cell)
            print("\""+"test_case_"+self.single_digit_add_zero(idx+1)+"_"+title_list[idx+2]+"\""+": [", file=f)
            print(formatted_cell, file=f)
            if idx != len(content_list)-1:
                print("],", file=f)
            else:
                print("]", file=f)
        f.close()

    def get_is_json(self, js):
        """
        to check if it is a json, very simple

        """
        try:
            json.loads(js)
        except ValueError:
            return False
        return True

    def get_is_number(self,s):
        """
        to check if it is a number, very simple

        """
        try:
            float(s)
            return True
        except ValueError:
            pass

        try:
            import unicodedata
            unicodedata.numeric(s)
            return True
        except (TypeError, ValueError):
            pass

        return False

    def str_custom_strip(self, input_str):
        """
        put custom strip and replace to string here
        not replacing spaces in the middle

        """
        output_str = input_str.strip().replace('\\n','').replace('\\','').replace('\\','').replace("“","\"").replace("”","\"")
        return output_str

    def strip_title_by_col_name(self, col_name):
        """
        after get col id, in this def we will strip test title

        response is a list
        """
        col_id = self.get_col_id_by_name(col_name)
        title_list = []

        for i in range(1, self.col_length):
            raw = self.ws.cell(row=i, column=col_id)
            if not raw.value:
                raw.value = "Empty test case name"
            r = raw.value.strip()
            r = r.replace("\n","_").replace(": ","_").replace(":","_").replace(" ", "_").replace(",","_")
            r = r.replace("__","_")
            title_list.append(r)

        for i in range(0,len(title_list)):
            print("[title]: "+str(i)+" =====" + title_list[i])
        return title_list

    def get_col_length_by_col_id(self, col_name):
        """
        the solution here is to get col id from col name, and then we are trying get the length of col

        in case of excel provider likes empty row, if 3 empty lines in a row, we define it hits bottom.

        better use test title to do this job, response is a list.
        """
        col_id = self.get_col_id_by_name(col_name)
        i = 2
        while (self.ws.cell(row=i, column=col_id).value) or (self.ws.cell(row=i+1, column=col_id).value) or (self.ws.cell(row=i+2, column=col_id).value):
            i = i + 1
        return i

    def single_digit_add_zero(self, num):
        if int(num) < 10:
            return "0"+str(num)
        else:
            return str(num)

class LoopExcelWsWb():

    def get_ws_list(self):
        """
        get .xlsx file list from current path
        """
        path = os.getcwd()
        excel_list = []

        for root, dirs, files in os.walk(path):
            for f in files:
                if ".xlsx" in f:
                    excel_list.append(f)

        return excel_list

    def get_wb_list(self, ws_name):
        """
        get worksheet list from input .xlsx file
        """
        ws = openpyxl.load_workbook(ws_name)
        return ws.worksheets

    def get_ws_and_wb_list(self):
        """
        get a 2-dimension [ws, wb] list 
        """
        ws_list = self.get_ws_list()
        ws_and_wb_list=[]
        for ws_name in ws_list:
            ws = openpyxl.load_workbook(ws_name)
            for wb in ws.worksheets:
                wb_name = wb.title
                ws_and_wb_list.append([ws_name,wb_name])
        print (ws_and_wb_list)
        return ws_and_wb_list

    def get_col_id_by_name(self, ws, col_name):
        """
        input column name, i.e. 'Expected Result', to output col id for further operations
        """
        for index in range(1, 100):
            row = ws.cell(row=1, column=index)
            if row.value and col_name in row.value:
                return index
        return False

if __name__ == "__main__":
    l = LoopExcelWsWb()
    blob_list = l.get_ws_and_wb_list()

    for w in blob_list:
        ws_name = w[0]
        wb_name = w[1]
        response_data_col = 'Expected Result'
        title_col = 'Test Name'

        wb = openpyxl.load_workbook(ws_name)
        ws = wb[wb_name]

        if l.get_col_id_by_name(ws,title_col) is not False:
            t = TransferAPIFromExcel(ws_name, wb_name, title_col)

            # get title list
            title_blob = t.strip_title_by_col_name(title_col)

            # data input to template
            response_blob = t.strip_response_data_by_col_name(response_data_col)
            r = t.fill_in_template_response(response_blob,'case_test_data.json')

            # output to file
            ws_name_strip = ws_name.strip(".xlsx").replace(" ","_")
            wb_name_strip = wb_name.replace(" ", "_")
            t.output_to_file(title_blob, r, ws_name_strip +"_"+ wb_name_strip + '_case_test_data.json')










